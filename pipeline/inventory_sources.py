# -*- coding: utf-8 -*-
"""采购流水线库存数据源。

所有适配器输出同一零件结构，库存审核无需知道数据来自 PartDB、ERP API、MCP
或 Excel。普通 ``stock`` 列默认视为未确认库存；只有明确映射
``confirmed_stock``，或设置 ``stock_is_confirmed: true``，才会用于生产承诺。

API/MCP 适配器只负责连接、分页和字段映射，不把任何客户 URL、Token 或账套
写入代码。客户配置放在被忽略的 config.yaml。
"""
import csv
import json
import os
import queue
import re
import shlex
import subprocess
import threading

import requests

from core import PartDB, SKILL_DIR, die, part_price, split_stock


ALIASES = {
    "ipn": ["IPN", "内部料号", "物料编码", "料号", "零件编号", "ID"],
    "model": ["型号", "物料型号", "规格型号", "物料名称", "名称", "NAME", "MPN"],
    "footprint": ["封装", "规格", "FOOTPRINT", "PACKAGE"],
    "confirmed_stock": ["已确认库存", "可用库存", "可承诺库存", "CONFIRMED_STOCK"],
    "unconfirmed_stock": ["未确认库存", "待盘点库存", "UNCONFIRMED_STOCK"],
    "stock": ["库存", "库存数量", "即时库存", "现存量", "可用量", "STOCK", "QTY"],
    "location": ["仓位", "库位", "库存位置", "仓库", "LOCATION"],
    "unit_price": ["单价", "含税单价", "采购单价", "UNIT_PRICE", "PRICE"],
    "supplier": ["供应商", "供应商名称", "SUPPLIER"],
}


def _norm(value):
    return re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", str(value or "").upper())


def _number(value):
    try:
        return float(str(value or "0").replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _path_get(obj, path, default=None):
    """读取 ``a.b[0].c`` 形式的响应字段路径。"""
    if not path:
        return obj
    cur = obj
    for part in str(path).replace("[", ".").replace("]", "").split("."):
        if not part:
            continue
        if isinstance(cur, dict):
            cur = cur.get(part, default)
        elif isinstance(cur, list) and part.isdigit() and int(part) < len(cur):
            cur = cur[int(part)]
        else:
            return default
        if cur is None:
            return default
    return cur


def _env(value):
    """支持 ${ENV_NAME}，避免把 API 凭证写进配置文件。"""
    if isinstance(value, str):
        return re.sub(r"\$\{([^}]+)\}", lambda m: os.getenv(m.group(1), ""), value)
    if isinstance(value, dict):
        return {k: _env(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_env(v) for v in value]
    return value


def _find_column(headers, logical, configured):
    wanted = configured.get(logical)
    if wanted:
        for header in headers:
            if _norm(header) == _norm(wanted):
                return header
        die(f"库存文件配置列 {logical}={wanted} 不存在；现有列：{headers}")
    for alias in ALIASES[logical]:
        for header in headers:
            if _norm(header) == _norm(alias):
                return header
    return None


def _merge_parts(rows, source_name, stock_is_confirmed=False):
    """规范化 API/MCP 行并按料号或型号合并重复库存记录。"""
    if not isinstance(rows, list):
        die(f"{source_name} 返回的数据不是列表")
    grouped = {}
    for index, raw in enumerate(rows, 1):
        if not isinstance(raw, dict):
            continue
        fields = raw.get("_fields") or {}
        def value(key, default=""):
            path = fields.get(key) or key
            result = _path_get(raw, path, default)
            return default if result is None else result

        ipn = str(value("ipn") or "").strip()
        model = str(value("model") or "").strip()
        key = _norm(ipn or model)
        if not key:
            continue
        generic = _number(value("stock", 0))
        confirmed = _number(value("confirmed_stock", 0))
        unconfirmed = _number(value("unconfirmed_stock", 0))
        if generic:
            if stock_is_confirmed:
                confirmed += generic
            else:
                unconfirmed += generic
        price = _number(value("unit_price", 0))
        location = str(value("location") or "").strip()
        supplier = str(value("supplier") or "").strip()
        part = grouped.setdefault(key, {
            "id": str(value("id") or f"{source_name}:{index}"),
            "ipn": ipn, "name": model,
            "footprint": str(value("footprint") or "").strip(),
            "confirmed_stock": 0.0, "unconfirmed_stock": 0.0,
            "locations": [], "unit_price": None, "supplier": "",
            "supplier_part": "", "note": "",
        })
        if ipn and not part["ipn"]:
            part["ipn"] = ipn
        if model and not part["name"]:
            part["name"] = model
        if location:
            part["locations"].append(location)
        part["confirmed_stock"] += confirmed
        part["unconfirmed_stock"] += unconfirmed
        if price > 0 and (part["unit_price"] is None or price < part["unit_price"]):
            part["unit_price"] = price
            part["supplier"] = supplier
    for part in grouped.values():
        part["locations"] = sorted(set(part["locations"]))
        if not stock_is_confirmed:
            part["note"] = "API/MCP通用库存默认列为未确认；人工盘点后方可承诺"
    return list(grouped.values())


class InventorySource:
    name = "inventory"

    def load_parts(self):
        raise NotImplementedError

    def enrich(self, part):
        return part


class PartDBSource(InventorySource):
    name = "PartDB"

    def __init__(self, cfg):
        self.db = PartDB(cfg)

    def load_parts(self):
        return [{"id": item.get("id"), "ipn": item.get("ipn") or "",
                 "name": item.get("name") or ""}
                for item in self.db.all_parts()]

    def enrich(self, part):
        detail, lots = self.db.part_lots(part["id"])
        confirmed, unconfirmed, locations = split_stock(lots)
        note = ""
        if not lots and detail.get("total_instock"):
            unconfirmed = _number(detail.get("total_instock"))
            note = "详情无批次明细；total_instock仅列为未确认"
        price, supplier, supplier_part = part_price(detail)
        footprint = detail.get("footprint") or {}
        return {
            "id": detail.get("id", part.get("id")),
            "ipn": detail.get("ipn") or part.get("ipn") or "",
            "name": detail.get("name") or part.get("name") or "",
            "footprint": footprint.get("name", "") if isinstance(footprint, dict) else str(footprint),
            "confirmed_stock": confirmed, "unconfirmed_stock": unconfirmed,
            "locations": locations, "unit_price": price,
            "supplier": supplier or "", "supplier_part": supplier_part or "",
            "note": note,
        }


class FileInventorySource(InventorySource):
    name = "Excel/CSV"

    def __init__(self, cfg):
        inv = (cfg or {}).get("inventory") or {}
        self.options = inv.get("file") or {}
        path = self.options.get("path")
        if not path:
            die("inventory.file.path 未配置库存 Excel/CSV 路径")
        self.path = path if os.path.isabs(path) else os.path.join(SKILL_DIR, path)
        if not os.path.exists(self.path):
            die(f"库存文件不存在：{self.path}")

    def _rows(self):
        ext = os.path.splitext(self.path)[1].lower()
        if ext in (".csv", ".tsv"):
            delimiter = "\t" if ext == ".tsv" else ","
            with open(self.path, encoding="utf-8-sig", newline="") as f:
                return list(csv.DictReader(f, delimiter=delimiter))
        if ext in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(self.path, data_only=True, read_only=True)
            sheet = self.options.get("sheet")
            ws = wb[sheet] if sheet else wb.active
            values = list(ws.iter_rows(values_only=True))
            header_row = max(1, int(self.options.get("header_row") or 1))
            if len(values) < header_row:
                return []
            headers = [str(x or "").strip() for x in values[header_row - 1]]
            return [dict(zip(headers, row)) for row in values[header_row:]
                    if any(x not in (None, "") for x in row)]
        die(f"库存文件暂只支持 CSV/TSV/XLSX/XLSM：{self.path}")

    def load_parts(self):
        rows = self._rows()
        if not rows:
            die(f"库存文件没有数据：{self.path}")
        headers = list(rows[0])
        configured = self.options.get("columns") or {}
        cols = {key: _find_column(headers, key, configured) for key in ALIASES}
        if not cols["ipn"] and not cols["model"]:
            die(f"库存文件找不到 IPN/内部料号或型号列；现有列：{headers}")
        if not any(cols[k] for k in ("confirmed_stock", "unconfirmed_stock", "stock")):
            die(f"库存文件找不到库存数量列；现有列：{headers}")
        mapped = []
        for row in rows:
            item = {"_fields": {key: cols[key] for key in cols if cols[key]}}
            item.update(row)
            mapped.append(item)
        return _merge_parts(mapped, self.name, bool(self.options.get("stock_is_confirmed", False)))


class ApiInventorySource(InventorySource):
    """通用 REST/HTTP 库存适配器，适配客户 ERP 的公开或私有 API。"""
    name = "HTTP API"

    def __init__(self, cfg):
        self.options = _env(((cfg or {}).get("inventory") or {}).get("api") or {})
        self.base_url = str(self.options.get("base_url") or "").rstrip("/")
        self.list_path = self.options.get("list_path") or "/inventory"
        if not self.base_url:
            die("inventory.api.base_url 未配置")
        self.timeout = int(self.options.get("timeout") or 40)
        self.fields = self.options.get("fields") or {}
        self.stock_is_confirmed = bool(self.options.get("stock_is_confirmed", False))
        self.session = requests.Session()
        self._login_if_configured()

    def _login_if_configured(self):
        login = self.options.get("login") or {}
        if not login:
            return
        path = login.get("path")
        if not path:
            die("inventory.api.login.path 未配置")
        url = self.base_url + (path if str(path).startswith("/") else "/" + str(path))
        method = str(login.get("method") or "POST").upper()
        headers = login.get("headers") or {}
        body = _env(login.get("body") or {})
        response = self.session.request(method, url, headers=headers, json=body, timeout=self.timeout)
        response.raise_for_status()
        token_path = login.get("token_path")
        if token_path:
            token = _path_get(response.json(), token_path)
            if not token:
                die("inventory.api.login.token_path 未从登录响应取到 Token")
            header = login.get("token_header") or "Authorization"
            prefix = login.get("token_prefix")
            self.session.headers[header] = f"{prefix} {token}" if prefix else str(token)

    def _request(self, params):
        method = str(self.options.get("method") or "GET").upper()
        headers = self.options.get("headers") or {}
        auth = self.options.get("auth") or {}
        if auth.get("type") == "bearer" and auth.get("token"):
            headers.setdefault("Authorization", f"Bearer {auth['token']}")
        elif auth.get("type") == "api_key" and auth.get("key") and auth.get("value"):
            headers.setdefault(auth["key"], auth["value"])
        url = self.base_url + (self.list_path if str(self.list_path).startswith("/") else "/" + str(self.list_path))
        kwargs = {"headers": headers, "timeout": self.timeout}
        if method == "GET":
            kwargs["params"] = params
        else:
            kwargs["json"] = dict(self.options.get("body") or {}, **params)
        response = self.session.request(method, url, **kwargs)
        response.raise_for_status()
        return response.json()

    def load_parts(self):
        pagination = self.options.get("pagination") or {}
        kind = str(pagination.get("type") or "none").lower()
        rows, page = [], int(pagination.get("start") or 1)
        for _ in range(int(pagination.get("max_pages") or 100)):
            params = dict(self.options.get("params") or {})
            if kind == "page":
                params[str(pagination.get("page_param") or "page")] = page
                params[str(pagination.get("size_param") or "page_size")] = int(pagination.get("size") or 100)
            elif kind == "offset":
                params[str(pagination.get("offset_param") or "offset")] = len(rows)
                params[str(pagination.get("size_param") or "limit")] = int(pagination.get("size") or 100)
            payload = self._request(params)
            batch = _path_get(payload, self.options.get("data_path"), payload)
            if isinstance(batch, dict):
                batch = batch.get("items") or batch.get("data") or batch.get("rows") or []
            if not isinstance(batch, list):
                die("inventory.api.data_path 未指向列表")
            rows.extend(batch)
            if kind == "none" or not batch or len(batch) < int(pagination.get("size") or 100):
                break
            page += 1
        return _merge_parts([{**row, "_fields": self.fields} for row in rows if isinstance(row, dict)],
                            self.name, self.stock_is_confirmed)


class McpInventorySource(InventorySource):
    """标准 MCP stdio 客户端：initialize 后调用配置中的库存工具。"""
    name = "MCP"

    def __init__(self, cfg):
        self.options = _env(((cfg or {}).get("inventory") or {}).get("mcp") or {})
        command = self.options.get("command")
        if not command:
            die("inventory.mcp.command 未配置")
        self.command = command if isinstance(command, list) else shlex.split(str(command), posix=False)
        self.tool = self.options.get("tool") or "inventory.list"
        self.fields = self.options.get("fields") or {}
        self.stock_is_confirmed = bool(self.options.get("stock_is_confirmed", False))
        self.timeout = int(self.options.get("timeout") or 60)

    def _rpc(self, proc, request_id, method, params=None):
        proc.stdin.write(json.dumps({"jsonrpc": "2.0", "id": request_id,
                                     "method": method, "params": params or {}}, ensure_ascii=False) + "\n")
        proc.stdin.flush()
        if not hasattr(proc, "_wb_lines"):
            proc._wb_lines = queue.Queue()

            def read_stdout():
                while True:
                    line = proc.stdout.readline()
                    proc._wb_lines.put(line)
                    if not line:
                        return

            threading.Thread(target=read_stdout, daemon=True).start()
        while True:
            try:
                line = proc._wb_lines.get(timeout=self.timeout)
            except queue.Empty:
                raise RuntimeError("MCP 服务响应超时")
            if not line:
                raise RuntimeError("MCP 服务提前退出")
            message = json.loads(line)
            if message.get("id") != request_id:
                continue
            if message.get("error"):
                raise RuntimeError(str(message["error"]))
            return message.get("result") or {}

    def load_parts(self):
        try:
            proc = subprocess.Popen(self.command, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                                    stderr=subprocess.PIPE, text=True, encoding="utf-8")
            self._rpc(proc, 1, "initialize", {"protocolVersion": "2024-11-05",
                                                "capabilities": {},
                                                "clientInfo": {"name": "seatable-production", "version": "1"}})
            proc.stdin.write(json.dumps({"jsonrpc": "2.0", "method": "notifications/initialized", "params": {}}) + "\n")
            proc.stdin.flush()
            result = self._rpc(proc, 2, "tools/call", {"name": self.tool,
                "arguments": self.options.get("arguments") or {}})
            payload = result.get("structuredContent") or result.get("content") or result
            if isinstance(payload, list):
                texts = [x.get("text") for x in payload if isinstance(x, dict) and x.get("text")]
                if len(texts) == 1:
                    try:
                        payload = json.loads(texts[0])
                    except (TypeError, ValueError):
                        payload = texts
            rows = _path_get(payload, self.options.get("data_path"), payload)
            if isinstance(rows, dict):
                rows = rows.get("items") or rows.get("data") or rows.get("rows") or []
            return _merge_parts([{**row, "_fields": self.fields} for row in rows if isinstance(row, dict)],
                                self.name, self.stock_is_confirmed)
        except (OSError, ValueError, RuntimeError) as exc:
            die(f"MCP 库存工具调用失败：{exc}")
        finally:
            try:
                proc.kill()
            except (UnboundLocalError, AttributeError, OSError):
                pass


def get_inventory_source(cfg):
    inv = (cfg or {}).get("inventory") or {}
    source = str(inv.get("source") or "").strip().lower()
    if not source and ((cfg or {}).get("partdb") or {}).get("enabled"):
        source = "partdb"
    if source == "partdb":
        return PartDBSource(cfg)
    if source in ("file", "excel", "csv", "erp-export", "kingdee-export", "jiandaoyun-export", "zentao-export"):
        return FileInventorySource(cfg)
    if source in ("api", "http", "rest", "kingdee-api", "jiandaoyun-api", "zentao-api"):
        return ApiInventorySource(cfg)
    if source in ("mcp", "mcp-tool"):
        return McpInventorySource(cfg)
    if not source:
        die("未配置库存源。请设置 inventory.source=partdb、api、mcp 或 file")
    die(f"不支持的库存源 inventory.source={source}；当前支持 partdb、api、mcp、file")
