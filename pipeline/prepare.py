# -*- coding: utf-8 -*-
"""合同与 BOM 解析。

支持：
- 合同 PDF：提取产品版本与采购数量。
- BOM CSV/XLSX：识别 IPN/型号/封装/单板用量列。
- 多版本 BOM 按各自套数扩量，同 IPN 合并求和。

合同未能可靠识别数量时不猜，输出待补 products.json 供人工修正后继续。
"""
import csv
import os
import re

from core import die, load_rules, save, save_csv


def norm(s):
    return re.sub(r"\s+", "", str(s or "")).upper()


def pdf_text(path):
    from pypdf import PdfReader
    r = PdfReader(path)
    return "\n".join((p.extract_text() or "") for p in r.pages)


def parse_contract(path, rules):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        text = pdf_text(path)
    elif ext in (".txt", ".md"):
        text = open(path, encoding="utf-8").read()
    else:
        die(f"合同暂只支持 PDF/TXT/MD：{path}")
    found = {}
    for pat in rules.get("contract_quantity_patterns", []):
        for m in re.finditer(pat, text, re.I):
            name = m.group("name").strip()
            qty = int(m.group("qty"))
            if qty > 0:
                found[name] = max(found.get(name, 0), qty)
    return text, [{"name": k, "qty": v} for k, v in found.items()]


def read_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    vals = list(ws.iter_rows(values_only=True))
    if not vals:
        return []
    header = [str(x or "").strip() for x in vals[0]]
    return [dict(zip(header, row)) for row in vals[1:] if any(x not in (None, "") for x in row)]


def read_bom(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return read_xlsx(path)
    if ext in (".csv", ".tsv"):
        enc = "utf-8-sig"
        delim = "\t" if ext == ".tsv" else ","
        with open(path, encoding=enc, newline="") as f:
            return list(csv.DictReader(f, delimiter=delim))
    die(f"BOM 暂只支持 CSV/TSV/XLSX：{path}")


ALIASES = {
    "ipn": ["IPN", "内部料号", "物料编码", "料号", "零件编号", "PARTDB"],
    "model": ["型号", "物料型号", "规格型号", "PART NUMBER", "MPN", "NAME", "物料名称"],
    "footprint": ["封装", "FOOTPRINT", "PACKAGE", "规格"],
    "qty": ["数量", "用量", "单板数量", "QTY", "QUANTITY", "PCS"],
    "description": ["描述", "品名", "物料描述", "DESCRIPTION", "类型"],
}


def find_col(row, logical):
    keys = list(row)
    for alias in ALIASES[logical]:
        na = norm(alias)
        for k in keys:
            if norm(k) == na:
                return k
    for alias in ALIASES[logical]:
        na = norm(alias)
        for k in keys:
            if na in norm(k):
                return k
    return None


def number(v):
    s = str(v or "").strip().replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else 0.0


def normalize_bom(rows, source, multiplier):
    if not rows:
        return []
    cols = {k: find_col(rows[0], k) for k in ALIASES}
    if not cols["ipn"] and not cols["model"]:
        die(f"BOM {source} 找不到 IPN 或型号列；现有列：{list(rows[0])}")
    if not cols["qty"]:
        die(f"BOM {source} 找不到数量/用量列；现有列：{list(rows[0])}")
    out = []
    for r in rows:
        q = number(r.get(cols["qty"]))
        if q <= 0:
            continue
        ipn = str(r.get(cols["ipn"]) or "").strip() if cols["ipn"] else ""
        model = str(r.get(cols["model"]) or "").strip() if cols["model"] else ""
        key = norm(ipn or model)
        if not key:
            continue
        out.append({
            "key": key, "ipn": ipn, "model": model,
            "footprint": str(r.get(cols["footprint"]) or "").strip() if cols["footprint"] else "",
            "description": str(r.get(cols["description"]) or "").strip() if cols["description"] else "",
            "unit_qty": q, "multiplier": multiplier,
            "need": q * multiplier, "source": os.path.basename(source),
        })
    return out


def merge_boms(items):
    merged = {}
    for x in items:
        key = x["key"]
        if key not in merged:
            merged[key] = {k: x[k] for k in ("ipn", "model", "footprint", "description")}
            merged[key].update({"need": 0.0, "sources": []})
        m = merged[key]
        m["need"] += x["need"]
        m["sources"].append({"bom": x["source"], "unit_qty": x["unit_qty"],
                             "multiplier": x["multiplier"], "need": x["need"]})
    return sorted(merged.values(), key=lambda x: (x["ipn"], x["model"]))


def configured_boms(products, rules):
    """合同识别出的产品自动映射 rules.yaml 的标准 BOM。"""
    configured = {p.get("name"): p.get("bom") for p in rules.get("products", [])}
    specs, missing = [], []
    for p in products:
        path = configured.get(p["name"])
        if path:
            if not os.path.isabs(path):
                path = os.path.join(os.path.dirname(os.path.abspath(__file__)), path)
            if os.path.exists(path):
                specs.append(f"{p['name']}={path}")
                continue
        missing.append(p["name"])
    if missing:
        die("以下合同产品未配置标准 BOM：" + "、".join(missing)
            + "。请在 pipeline/rules.yaml 的 products[].bom 填文件路径，"
              "或本次用 --bom 产品名=文件路径。")
    return specs


def run(run_id, contract, bom_specs):
    rules = load_rules()
    text, products = parse_contract(contract, rules)
    save(run_id, "contract.json", {"source": os.path.abspath(contract),
                                    "products": products, "text": text})
    if not products:
        die("合同未可靠识别产品数量。请检查合同文本/产品别名，不会猜测数量。")
    if not bom_specs:
        bom_specs = configured_boms(products, rules)
        print(f"[自动] 已按产品映射 {len(bom_specs)} 份标准 BOM")
    product_qty = {p["name"]: p["qty"] for p in products}
    all_items = []
    specs_out = []
    for spec in bom_specs:
        if "=" not in spec:
            die("--bom 格式必须是 产品名=文件路径，例如 无GPS版=bom.xlsx")
        name, path = spec.split("=", 1)
        qty = product_qty.get(name)
        if qty is None:
            die(f"合同未识别 {name} 数量；请用 --qty {name}=数量 覆盖，或修改 contract.json")
        rows = read_bom(path)
        normed = normalize_bom(rows, path, qty)
        all_items.extend(normed)
        specs_out.append({"name": name, "path": os.path.abspath(path), "qty": qty,
                          "bom_rows": len(normed)})
    merged = merge_boms(all_items)
    save(run_id, "prepared.json", {"products": specs_out, "materials": merged})
    save_csv(run_id, "合并备料表.csv", merged,
             ["ipn", "model", "footprint", "description", "need", "sources"])
    print(f"[完成] 合并 {len(merged)} 种物料，总备料量 {sum(x['need'] for x in merged):,.0f}")
    return merged
